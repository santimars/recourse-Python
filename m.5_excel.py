'''
Excel es un software de hoja de cálculo usado en todo el mundo.

Sus inconvenientes son que no puede trabajar con cantidades de datos 
exageradamente grandes, aunque nos permite realizar multitud de operaciones, 
no nos permite aplicar una lógica de programación, su rendimiento decae con el 
volumen de datos y sus ficheros (.xls) tienen un formato muy complejo y 
visual, lo que hace que sean muy pesados.

Por ejemplo, una tabla de Excel con 1 millón de registros ocupa,
 aproximadamente, un tamaño de unos 12MB.

En programación podremos trabajar con ficheros Excel y suplir 
algunas de sus debilidades, o trabajaremos más eficientemente con ficheros CSV.

Vamos a ver cómo trabajar con ficheros tipo Excel (.xls),
 para ello utilizaremos la librería openpyxl, que nos permite 
 acceder y modificar ficheros Excel. Podemos encontrar 
 la documentación en https://openpyxl.readthedocs.io/en/stable/.
   Lo primero será importar dicha librería a nuestro programa mediante
     la instrucción:
'''

import openpyxl

# Abrir Archivo Excel (workbook) # importante poner siempre la ruta de acceso como esta el archibo guardado
wb = openpyxl.load_workbook("/Users/santirodriguez/Desktop/02_Excel_data.xlsx")

# imprimir los nombres
print("Nombre de hojas: ")
print(wb.sheetnames)

print("\nNombre de hojas:")
for sheet in wb:
  print("-",sheet.title)

#hoja 1

#crea una hoja que haga referencia a la primera hoja de excel

hoja_uno = wb.sheetnames[0]
print("\n Primera Hoja:")
print("-",hoja_uno)  

hoja_otros = wb.sheetnames[2]

#del hoja_otros
#for sheet in wb:
#  print("-",sheet.title)

wb.save("02_Excel_data.xlsx")  



'''
hay que invertir  mas tiempo leyendo la documentacion de las librerias y buscar las 
solucion y imprementarla en codigo
 un programador tiene 70% leyendo 30 % programando
incovenientes
No puede trabajar con cantidades de datos exageradamente grandes
nos permite realizar multitud de operaciones, pero no aplicar una 
logica de programacion
su rendimiento decae con el volumen de datos 
los ficheros excel .xls tienen un formato muy complejo y visual
lo que hace que sean muy pesados
en Programacion prodemos trabajar con ficheros excel y suplir alfunasd e sus debilidades o trabajaremos mas eficientemente 
con ficheros CSV.

'''
"""
IMPORTANTE

TENER CERRADO SIEMPRE EL EXCEL CUANDO VALLAMOS A EDITAR
ya que se puede corromper 
no se permite que este abierto el excel en cualquier lado
"""
# Hoja 1 Lectura
# Acceso a una celda de una hoja concreta


# Crear una variable que haga referencia a la primera hoja del excel
hoja_uno = wb["address"]

# Acceder a una celda directamente

print(hoja_uno["A1"].value) # contenido
print(hoja_uno["B1"])       # Referencia

#   Acceder a una  celda a traves de una nomenclatura fila-columna

print(hoja_uno.cell(row=1,column=2).value) # con .value podemos leer su contenido que esta en las  celdas en excel
# mejor forma de acceder a los valores de las columnas


""                                   """Acceso a multiples celdas de una hoja concreta """
# Crear una variable que haga regerencia a la primera hoja de excel
multiple_cells = hoja_uno["A1":"F3"]
for row in multiple_cells:
  for cell in row:
    print("-",cell.value)

''  '''Acceso a todas las filas de una hoja concreta'''


for fila in hoja_uno.rows:
  for columna in fila:
    print (columna.coordinate, columna.value)
  print("----------fin de la lista----------------")

# recorre todas las filas donde tenga datos 
  # coordinate te devuelve la coordenada de la fila que tenga el dato guardado

'''
Acabamos de hacer una actualizacion de nuestra hoja de python 
pero python no tiene la capacidad de leer los nuevos cambios
que hemos hecho.
ya que hace una especie de conexion con el excel del momento en el que le 
digimos que buscara.

hay que decirle a python las actualizaciones que hemos hecho

nesecitamos refresacar el fichero
'''

wb = openpyxl.load_workbook("/Users/santirodriguez/Desktop/02_Excel_data.xlsx") # aqui le referescamos la informacion de nuevo
hoja_uno = wb["address"]

for fila in hoja_uno.rows:
  for columna in fila:
     if columna.value != None:
      print (columna.coordinate, columna.value)
  print("----------filas vacias ----------------")
   

""                    """Acceso a todas las columnas de una hoja concreta"""

for columna in hoja_uno.columns:
  for fila in columna:

     print(fila.coordinate,fila.value)
  print("----COLUMNA VACIA----")  


##################################################################################################################################
wb = openpyxl.load_workbook("/Users/santirodriguez/Desktop/02_Excel_data.xlsx") # aqui le referescamos la informacion de nuevo
hoja_uno = wb["address"]

'HOJA 2'
'Modificacion de celdas de una hoja concreta'

import datetime

# Crear una variavble que haga referencia a la segunda hoja del excel
hoja_dos = wb["ventas"]

# modificar una celda (3 formas diferentes)

hoja_dos["B6"] = 5
hoja_dos["C6"].value = 5 
hoja_dos.cell(row = 6,column = 4,value = 5)

#Añadir la hora actual en la celda E2

hoja_dos["E2"].value =datetime.datetime.now()


#Guardar cambios (IMPORTANTE! El excel deve estar cerrado)

wb.save("/Users/santirodriguez/Desktop/02_Excel_data.xlsx") # Guardamos datos antes de ejecutar el resto
 
# Modificar la celda B2, que contiene un 40, por un 99
print("Celda B2 antes de modificarla:",hoja_dos["B2"].value)
hoja_dos["B2"].value = 99 # modificando la celda B2 por 99 
print("Celda B2 despues de modificarla: ", hoja_dos["B2"].value)

#Modificar la celda C2, pero eta vez conuna formula
print("Celda C2 antes de su modificacion ",hoja_dos["C2"].value)
x = 2
y = 3
f = "=SUM(B{},{})".format(x,y)
hoja_dos["C2"] = f # las funciones que debemos usar en excel deber ser de la forma inglesa no capta en español
print("Celda C2 después de su modificación: ",hoja_dos["C2"].value)

#Guardar cambios (IMPORTANTE! El excel deve estar cerrado)

wb.save("/Users/santirodriguez/Desktop/02_Excel_data.xlsx") # Guardamos datos antes de ejecutar el resto

# sabiendo esto podemos hacer una funcion donde podemos automatizar el valor sumar
def sumar(a,b,c):
  return "=SUM({}{},{})".format(a,b,c)
    
hoja_dos["C3"] = sumar("B",3,3)
print ("La fórmula de Celda C3 es :",hoja_dos["C3"].value)
hoja_dos["C4"]  = sumar("B",4,3)
print("la formula de la celda C4 es : ",hoja_dos["C4"].value)

#Guardar cambios (IMPORTANTE! El excel deve estar cerrado)

wb.save("/Users/santirodriguez/Desktop/02_Excel_data.xlsx") # Guardamos datos antes de ejecutar el resto

''                       '''Añadir contenido nuevo a una hoja concreta'''


hoja_dos.append(["Junio",99,100])
# asi agregamos grandes conjuntos de datos en excel 

#Guardar cambios
wb.save("/Users/santirodriguez/Desktop/02_Excel_data.xlsx") # Guardamos datos antes de ejecutar el resto

''                    '''Crear una hoja nueva y darle nombre'''



# creamos una nueva hoja(por defecto al final, si no,
# le podemos pasar como parametro el indice de la posicion que se quiera).

hoja_otros = wb.create_chartsheet("Otros") 

#Imprimir los nombres de las hojas 
print("Nombre de las hojas: ")
print(wb.sheetnames)
#Guardar cambios
wb.save("/Users/santirodriguez/Desktop/02_Excel_data.xlsx") 

''                                       '''AVANZADO. Generar graficas'''
from openpyxl.chart import (
    AreaChart,
    Reference,
    Series,
)

#todo lo que es diseño como graficas y tablas con openpyxl podemos hacerlos 

#Abrir un archivo Excel(workbook)

# Definimos la hoja dos (hojas de ventas)
hoja_dos = wb["ventas"]

# configuracion y creacion de la grafica
grafica = AreaChart()
"grafica = openpyxl.chart.AreaChart()"

# Para versiones nuevas de Anaconda : grafica = AreaChart()

grafica.title = "Grafica de Area"
grafica.style = 13
grafica.x_axis.title = "Periodo"
grafica.y_axis.title = "Utilidades"

"periodo = openpyxl.chart.Reference(hoja_dos, min_col =1, min_row = 2, max_col = 1,max_row = 7)"
periodo = Reference(hoja_dos, min_col =1, min_row = 2, max_col = 1, max_row = 7)
# Para versiones nuevas de Anaconda: periodo = Reference(hoja_dos, min_col =1, min_row = 2, max_col = 1, max_row = 7)
"utilidades = openpyxl.chart.Reference(hoja_dos, min_col = 2, min_row = 1, max_col = 3, max_row = 7)"
utilidades = Reference(hoja_dos, min_col = 2, min_row = 1, max_col = 3, max_row = 7)
# Para versiones nuevas de Anaconda: Utilidades = Reference(hoja_dos, min_col = 2, min_row = 1, max_col = 3, max_row = 7)

#estos metodos agregamos el metodo de las x y el otro el metodo de la y
grafica.add_data(utilidades,titles_from_data = True)
grafica.set_categories(periodo)

# le decimo en que celda va estar el grafico
hoja_dos.add_chart(grafica, "A1O")


#Guardar cambios
wb.save("/Users/santirodriguez/Desktop/02_Excel_data.xlsx") # Guardamos datos antes de ejecutar el resto
#Creamos una tabla en excel para mostrar los valores de utilidades por
#periodes
